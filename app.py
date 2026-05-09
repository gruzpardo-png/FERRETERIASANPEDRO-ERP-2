
import os
import json
from datetime import datetime, date
from zoneinfo import ZoneInfo
from functools import wraps
from io import BytesIO

from flask import Flask, request, redirect, url_for, session, flash, send_file, render_template_string
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename


APP_NAME = "Ferretería Cloud ERP"
APP_VERSION = "v1.0 Render MVP"
CHILE_TZ = ZoneInfo("America/Santiago")
IVA_RATE = float(os.environ.get("IVA_RATE", "0.19"))
SUPERIOR_USERNAME = os.environ.get("SUPERIOR_USERNAME", "gus").strip().lower()


def now_cl():
    return datetime.now(CHILE_TZ)


def now_str():
    return now_cl().strftime("%Y-%m-%d %H:%M:%S")


def today_str():
    return now_cl().date().isoformat()


def get_database_uri():
    database_url = os.environ.get("DATABASE_URL", "").strip()
    if database_url:
        if database_url.startswith("postgres://"):
            database_url = database_url.replace("postgres://", "postgresql://", 1)
        return database_url

    db_path = os.environ.get("DATABASE_PATH", "/data/ferreteria_cloud_erp.db")
    parent = os.path.dirname(os.path.abspath(db_path))
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)
    return "sqlite:///" + os.path.abspath(db_path)


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-en-render")
app.config["SQLALCHEMY_DATABASE_URI"] = get_database_uri()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024

db = SQLAlchemy(app)


# ============================================================
# MODELOS
# ============================================================

class Company(db.Model):
    __tablename__ = "erp_companies"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(180), nullable=False, default="Ferretería San Pedro")
    legal_name = db.Column(db.String(180), nullable=False, default="Socom Ruz Spa")
    rut = db.Column(db.String(30), nullable=False, default="77.351.584-0")
    address_1 = db.Column(db.String(220), default="Av. 21 de Mayo 73, Quillota")
    address_2 = db.Column(db.String(220), default="Av. 21 de Mayo 420, Quillota")
    whatsapp_1 = db.Column(db.String(40), default="+56 9 3563 0950")
    whatsapp_2 = db.Column(db.String(40), default="+56 9 4907 3615")
    created_at = db.Column(db.String(30), default=now_str)


class User(db.Model):
    __tablename__ = "erp_users"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("erp_companies.id"), nullable=False)
    username = db.Column(db.String(60), unique=True, nullable=False)
    full_name = db.Column(db.String(160), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(40), default="operador")
    permissions = db.Column(db.Text, default="{}")
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.String(30), default=now_str)
    updated_at = db.Column(db.String(30), default=now_str)


class AuditLog(db.Model):
    __tablename__ = "erp_audit_log"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer)
    username = db.Column(db.String(80))
    module = db.Column(db.String(80), nullable=False)
    action = db.Column(db.String(80), nullable=False)
    entity = db.Column(db.String(120))
    entity_id = db.Column(db.String(60))
    old_value = db.Column(db.Text)
    new_value = db.Column(db.Text)
    created_at = db.Column(db.String(30), default=now_str)


class Branch(db.Model):
    __tablename__ = "erp_branches"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    code = db.Column(db.String(30), nullable=False)
    name = db.Column(db.String(120), nullable=False)
    address = db.Column(db.String(220))
    active = db.Column(db.Boolean, default=True)


class Warehouse(db.Model):
    __tablename__ = "erp_warehouses"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    branch_id = db.Column(db.Integer, db.ForeignKey("erp_branches.id"), nullable=False)
    code = db.Column(db.String(30), nullable=False)
    name = db.Column(db.String(120), nullable=False)
    active = db.Column(db.Boolean, default=True)


class Product(db.Model):
    __tablename__ = "erp_products"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    sku = db.Column(db.String(80), nullable=False)
    barcode = db.Column(db.String(80))
    description = db.Column(db.String(420), nullable=False)
    category = db.Column(db.String(120))
    brand = db.Column(db.String(120))
    unit = db.Column(db.String(30), default="UN")
    cost_net = db.Column(db.Float, default=0)
    sale_price_gross = db.Column(db.Float, default=0)
    min_price_gross = db.Column(db.Float, default=0)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.String(30), default=now_str)
    updated_at = db.Column(db.String(30), default=now_str)
    __table_args__ = (db.UniqueConstraint("company_id", "sku", name="uq_erp_product_company_sku"),)


class StockBalance(db.Model):
    __tablename__ = "erp_stock_balances"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("erp_products.id"), nullable=False)
    branch_id = db.Column(db.Integer, db.ForeignKey("erp_branches.id"), nullable=False)
    warehouse_id = db.Column(db.Integer, db.ForeignKey("erp_warehouses.id"), nullable=False)
    physical = db.Column(db.Float, default=0)
    reserved = db.Column(db.Float, default=0)
    committed = db.Column(db.Float, default=0)
    available = db.Column(db.Float, default=0)
    updated_at = db.Column(db.String(30), default=now_str)
    product = db.relationship("Product")
    branch = db.relationship("Branch")
    warehouse = db.relationship("Warehouse")
    __table_args__ = (db.UniqueConstraint("company_id", "product_id", "warehouse_id", name="uq_erp_stock_product_warehouse"),)


class StockMovement(db.Model):
    __tablename__ = "erp_stock_movements"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("erp_products.id"), nullable=False)
    branch_id = db.Column(db.Integer, db.ForeignKey("erp_branches.id"), nullable=False)
    warehouse_id = db.Column(db.Integer, db.ForeignKey("erp_warehouses.id"), nullable=False)
    movement_type = db.Column(db.String(60), nullable=False)
    document_type = db.Column(db.String(60))
    document_number = db.Column(db.String(80))
    qty_in = db.Column(db.Float, default=0)
    qty_out = db.Column(db.Float, default=0)
    stock_after = db.Column(db.Float, default=0)
    cost_net = db.Column(db.Float, default=0)
    price_gross = db.Column(db.Float, default=0)
    user_id = db.Column(db.Integer)
    notes = db.Column(db.Text)
    created_at = db.Column(db.String(30), default=now_str)
    product = db.relationship("Product")
    branch = db.relationship("Branch")
    warehouse = db.relationship("Warehouse")


class Customer(db.Model):
    __tablename__ = "erp_customers"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    code = db.Column(db.String(80))
    rut = db.Column(db.String(40))
    name = db.Column(db.String(220), nullable=False)
    email = db.Column(db.String(160))
    phone = db.Column(db.String(80))
    address = db.Column(db.String(240))
    credit_limit = db.Column(db.Float, default=0)
    active = db.Column(db.Boolean, default=True)


class Supplier(db.Model):
    __tablename__ = "erp_suppliers"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    code = db.Column(db.String(80))
    rut = db.Column(db.String(40))
    name = db.Column(db.String(220), nullable=False)
    email = db.Column(db.String(160))
    phone = db.Column(db.String(80))
    address = db.Column(db.String(240))
    active = db.Column(db.Boolean, default=True)


class Sale(db.Model):
    __tablename__ = "erp_sales"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    branch_id = db.Column(db.Integer)
    warehouse_id = db.Column(db.Integer)
    customer_id = db.Column(db.Integer)
    user_id = db.Column(db.Integer)
    document_type = db.Column(db.String(60), nullable=False)
    document_number = db.Column(db.String(80), nullable=False)
    payment_method = db.Column(db.String(80))
    status = db.Column(db.String(40), default="Emitida")
    total_net = db.Column(db.Float, default=0)
    iva = db.Column(db.Float, default=0)
    total_gross = db.Column(db.Float, default=0)
    cost_total_net = db.Column(db.Float, default=0)
    contribution = db.Column(db.Float, default=0)
    margin_pct = db.Column(db.Float, default=0)
    created_at = db.Column(db.String(30), default=now_str)


class SaleLine(db.Model):
    __tablename__ = "erp_sale_lines"
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey("erp_sales.id"), nullable=False)
    product_id = db.Column(db.Integer)
    sku = db.Column(db.String(80))
    description = db.Column(db.String(420))
    qty = db.Column(db.Float, default=1)
    price_gross = db.Column(db.Float, default=0)
    discount_pct = db.Column(db.Float, default=0)
    final_price_gross = db.Column(db.Float, default=0)
    cost_net = db.Column(db.Float, default=0)
    total_gross = db.Column(db.Float, default=0)
    contribution = db.Column(db.Float, default=0)
    margin_pct = db.Column(db.Float, default=0)


class Purchase(db.Model):
    __tablename__ = "erp_purchases"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    supplier_id = db.Column(db.Integer)
    branch_id = db.Column(db.Integer)
    warehouse_id = db.Column(db.Integer)
    document_type = db.Column(db.String(60), nullable=False)
    document_number = db.Column(db.String(80), nullable=False)
    status = db.Column(db.String(40), default="Recepcionada")
    total_net = db.Column(db.Float, default=0)
    iva = db.Column(db.Float, default=0)
    total_gross = db.Column(db.Float, default=0)
    created_at = db.Column(db.String(30), default=now_str)


class PurchaseLine(db.Model):
    __tablename__ = "erp_purchase_lines"
    id = db.Column(db.Integer, primary_key=True)
    purchase_id = db.Column(db.Integer, db.ForeignKey("erp_purchases.id"), nullable=False)
    product_id = db.Column(db.Integer)
    sku = db.Column(db.String(80))
    description = db.Column(db.String(420))
    qty = db.Column(db.Float, default=1)
    cost_net = db.Column(db.Float, default=0)
    total_net = db.Column(db.Float, default=0)


class Quote(db.Model):
    __tablename__ = "erp_quotes"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    number = db.Column(db.String(80), nullable=False)
    customer_id = db.Column(db.Integer)
    customer_name = db.Column(db.String(220))
    contact = db.Column(db.String(160))
    user_id = db.Column(db.Integer)
    valid_days = db.Column(db.Integer, default=2)
    status = db.Column(db.String(40), default="Borrador")
    total_net = db.Column(db.Float, default=0)
    iva = db.Column(db.Float, default=0)
    total_gross = db.Column(db.Float, default=0)
    cost_total_net = db.Column(db.Float, default=0)
    contribution = db.Column(db.Float, default=0)
    margin_pct = db.Column(db.Float, default=0)
    created_at = db.Column(db.String(30), default=now_str)


class QuoteLine(db.Model):
    __tablename__ = "erp_quote_lines"
    id = db.Column(db.Integer, primary_key=True)
    quote_id = db.Column(db.Integer, db.ForeignKey("erp_quotes.id"), nullable=False)
    product_id = db.Column(db.Integer)
    sku = db.Column(db.String(80))
    description = db.Column(db.String(420))
    qty = db.Column(db.Float, default=1)
    price_gross = db.Column(db.Float, default=0)
    discount_pct = db.Column(db.Float, default=0)
    final_price_gross = db.Column(db.Float, default=0)
    cost_net = db.Column(db.Float, default=0)
    total_gross = db.Column(db.Float, default=0)
    contribution = db.Column(db.Float, default=0)
    margin_pct = db.Column(db.Float, default=0)


class CashSession(db.Model):
    __tablename__ = "erp_cash_sessions"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    branch_id = db.Column(db.Integer)
    user_id = db.Column(db.Integer)
    opening_amount = db.Column(db.Float, default=0)
    closing_amount = db.Column(db.Float, default=0)
    status = db.Column(db.String(40), default="Abierta")
    opened_at = db.Column(db.String(30), default=now_str)
    closed_at = db.Column(db.String(30))


class Delivery(db.Model):
    __tablename__ = "erp_deliveries"
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, nullable=False)
    sale_id = db.Column(db.Integer)
    document_type = db.Column(db.String(60), nullable=False)
    document_number = db.Column(db.String(80), nullable=False)
    status = db.Column(db.String(40), default="Pendiente")
    vehicle_plate = db.Column(db.String(40))
    driver = db.Column(db.String(160))
    helper = db.Column(db.String(160))
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer)
    created_at = db.Column(db.String(30), default=now_str)
    delivered_at = db.Column(db.String(30))
    __table_args__ = (db.UniqueConstraint("company_id", "document_type", "document_number", name="uq_erp_delivery_document"),)


class DeliveryLine(db.Model):
    __tablename__ = "erp_delivery_lines"
    id = db.Column(db.Integer, primary_key=True)
    delivery_id = db.Column(db.Integer, db.ForeignKey("erp_deliveries.id"), nullable=False)
    product_id = db.Column(db.Integer)
    sku = db.Column(db.String(80))
    description = db.Column(db.String(420))
    qty_document = db.Column(db.Float, default=0)
    qty_delivered = db.Column(db.Float, default=0)
    qty_pending = db.Column(db.Float, default=0)
    status = db.Column(db.String(40), default="Pendiente")


# ============================================================
# PERMISOS Y HELPERS
# ============================================================

PERMISSION_LABELS = {
    "dashboard": "Dashboard ejecutivo",
    "pos": "POS / Ventas rápidas",
    "quotes": "Cotizaciones",
    "products": "Productos y precios",
    "stock": "Stock inteligente / Kardex",
    "purchases": "Compras y proveedores",
    "customers": "Clientes / CRM",
    "suppliers": "Proveedores",
    "cash": "Caja y turnos",
    "deliveries": "Despachos / bodega",
    "dte": "Documentos DTE",
    "ai": "Asistente IA",
    "users": "Usuarios y permisos",
    "audit": "Auditoría",
}


def current_company():
    company = Company.query.first()
    return company


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return User.query.filter_by(id=uid, is_active=True).first()


def user_permissions(user=None):
    user = user or current_user()
    if not user:
        return {}
    if is_superior(user):
        return {"all": True, "superior": True}
    if user.role == "admin":
        return {"all": True}
    try:
        return json.loads(user.permissions or "{}")
    except Exception:
        return {}


def is_superior(user=None):
    user = user or current_user()
    return bool(user and user.username.lower() == SUPERIOR_USERNAME)


def is_admin(user=None):
    user = user or current_user()
    return bool(user and user.role == "admin")


def has_perm(module_name):
    user = current_user()
    if not user:
        return False
    if is_superior(user) or is_admin(user):
        return True
    return bool(user_permissions(user).get(module_name))


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def permission_required(module_name):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user():
                return redirect(url_for("login"))
            if not has_perm(module_name):
                flash("No tienes permiso para acceder a esa sección.", "error")
                return redirect(url_for("index"))
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def parse_float(value, default=0.0):
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("$", "").replace(" ", "").replace("\xa0", "")
    if not s:
        return default
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return default


def calc_line(qty, price_gross, discount_pct, cost_net):
    qty = parse_float(qty, 0)
    price_gross = parse_float(price_gross, 0)
    discount_pct = parse_float(discount_pct, 0)
    cost_net = parse_float(cost_net, 0)
    final_price = price_gross * (1 - discount_pct / 100)
    total_gross = final_price * qty
    sale_net_unit = final_price / (1 + IVA_RATE) if final_price else 0
    contribution = (sale_net_unit - cost_net) * qty
    margin = (final_price / (cost_net * (1 + IVA_RATE)) - 1) if cost_net > 0 else 0
    return {
        "final_price": final_price,
        "total_gross": total_gross,
        "contribution": contribution,
        "margin": margin,
    }


def write_audit(module, action, entity=None, entity_id=None, old="", new=""):
    user = current_user()
    company = current_company()
    log = AuditLog(
        company_id=company.id if company else 1,
        user_id=user.id if user else None,
        username=user.username if user else "sistema",
        module=module,
        action=action,
        entity=entity,
        entity_id=str(entity_id or ""),
        old_value=str(old or ""),
        new_value=str(new or ""),
    )
    db.session.add(log)
    db.session.commit()


def default_branch_and_warehouse():
    company = current_company()
    branch = Branch.query.filter_by(company_id=company.id, code="001").first()
    warehouse = Warehouse.query.filter_by(company_id=company.id, code="001").first()
    return branch, warehouse


def get_or_create_balance(product_id, branch_id, warehouse_id):
    company = current_company()
    balance = StockBalance.query.filter_by(
        company_id=company.id,
        product_id=product_id,
        warehouse_id=warehouse_id
    ).first()
    if not balance:
        balance = StockBalance(
            company_id=company.id,
            product_id=product_id,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            physical=0,
            reserved=0,
            committed=0,
            available=0,
        )
        db.session.add(balance)
        db.session.commit()
    return balance


def add_stock_movement(product, branch_id, warehouse_id, movement_type, qty_in=0, qty_out=0, document_type="", document_number="", price_gross=0, cost_net=None, notes=""):
    company = current_company()
    balance = get_or_create_balance(product.id, branch_id, warehouse_id)
    new_stock = float(balance.physical or 0) + float(qty_in or 0) - float(qty_out or 0)
    balance.physical = new_stock
    balance.available = new_stock - float(balance.reserved or 0) - float(balance.committed or 0)
    balance.updated_at = now_str()

    mv = StockMovement(
        company_id=company.id,
        product_id=product.id,
        branch_id=branch_id,
        warehouse_id=warehouse_id,
        movement_type=movement_type,
        document_type=document_type,
        document_number=document_number,
        qty_in=float(qty_in or 0),
        qty_out=float(qty_out or 0),
        stock_after=new_stock,
        cost_net=float(cost_net if cost_net is not None else product.cost_net or 0),
        price_gross=float(price_gross or 0),
        user_id=current_user().id if current_user() else None,
        notes=notes,
    )
    db.session.add(mv)
    db.session.commit()
    return mv


def product_by_sku(sku):
    company = current_company()
    return Product.query.filter_by(company_id=company.id, sku=str(sku).strip()).first()


@app.template_filter("money")
def money(value):
    try:
        n = float(value or 0)
    except Exception:
        n = 0
    return "$" + f"{n:,.0f}".replace(",", ".")


@app.template_filter("percent")
def percent(value):
    try:
        n = float(value or 0)
    except Exception:
        n = 0
    return f"{n*100:.1f}%".replace(".", ",")


# ============================================================
# UI BASE
# ============================================================

BASE_HTML = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>{{ title }} · {{ app_name }}</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
:root{
  --bg:#f4f7fb; --panel:#ffffff; --ink:#0f172a; --muted:#64748b;
  --brand:#0f766e; --blue:#2563eb; --nav:#07111f; --border:#dbe3ef;
  --danger:#b91c1c; --warn:#b45309; --ok:#047857;
  --shadow:0 14px 38px rgba(15,23,42,.08);
}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--ink);font-family:Inter,Segoe UI,Roboto,Arial,sans-serif}
a{text-decoration:none;color:inherit}
.topbar{background:linear-gradient(135deg,#07111f,#0f172a);color:white;padding:14px 22px;display:flex;align-items:center;gap:18px;position:sticky;top:0;z-index:9;box-shadow:0 10px 30px rgba(2,6,23,.25)}
.brand{min-width:220px}
.brand strong{display:block;font-size:18px}.brand span{display:block;font-size:12px;color:#cbd5e1}
.nav{display:flex;flex-wrap:wrap;gap:8px;align-items:center;flex:1}
.nav a{font-size:13px;font-weight:800;padding:9px 11px;border:1px solid rgba(255,255,255,.15);border-radius:999px;background:rgba(255,255,255,.05)}
.nav a:hover{background:rgba(20,184,166,.25);border-color:#2dd4bf}
.userbox{font-size:12px;color:#cbd5e1;text-align:right}
.layout{max-width:1480px;margin:0 auto;padding:24px}
.card{background:var(--panel);border:1px solid var(--border);border-radius:18px;padding:20px;box-shadow:var(--shadow);margin-bottom:18px}
h1,h2,h3{margin:0 0 12px} p{color:var(--muted)}
.grid{display:grid;gap:14px}.grid-2{grid-template-columns:repeat(2,1fr)}.grid-3{grid-template-columns:repeat(3,1fr)}.grid-4{grid-template-columns:repeat(4,1fr)}
.stat{background:white;border:1px solid var(--border);border-radius:16px;padding:16px;box-shadow:var(--shadow)}
.stat span{display:block;color:var(--muted);font-size:12px}.stat strong{font-size:23px}
label{display:block;font-size:12px;font-weight:800;margin-bottom:6px;color:#334155}
input,select,textarea{width:100%;border:1px solid #cbd5e1;border-radius:12px;padding:10px 11px;background:white;color:#111827}
textarea{min-height:90px}
.form-row{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.form-row-2{display:grid;grid-template-columns:repeat(2,1fr);gap:12px}.form-row-3{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.actions{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-top:12px}
.btn{border:0;border-radius:12px;padding:10px 13px;font-weight:900;cursor:pointer;display:inline-flex;align-items:center;justify-content:center}
.btn-primary{background:var(--brand);color:white}.btn-secondary{background:#e5e7eb;color:#111827}.btn-blue{background:var(--blue);color:white}.btn-danger{background:#fee2e2;color:#991b1b}.btn-small{padding:6px 8px;font-size:11px}
.table-wrap{overflow:auto}.table{width:100%;border-collapse:collapse;background:white;border:1px solid var(--border);border-radius:14px;overflow:hidden}
th,td{padding:10px;border-bottom:1px solid var(--border);font-size:13px;text-align:left;vertical-align:top}
th{background:#f8fafc;color:#475569;text-transform:uppercase;font-size:11px;letter-spacing:.04em}
.num{text-align:right;white-space:nowrap;font-weight:800}
.badge{display:inline-flex;border-radius:999px;padding:4px 9px;font-size:11px;font-weight:900;background:#e5e7eb}
.badge.ok{background:#d1fae5;color:#065f46}.badge.warn{background:#fef3c7;color:#92400e}.badge.bad{background:#fee2e2;color:#991b1b}.badge.info{background:#dbeafe;color:#1e40af}
.flash{padding:13px 15px;border-radius:14px;margin-bottom:14px;font-weight:800}.flash.success{background:#d1fae5;color:#065f46}.flash.error{background:#fee2e2;color:#991b1b}.flash.info{background:#dbeafe;color:#1e40af}
.login-bg{min-height:100vh;display:flex;align-items:center;justify-content:center;background:radial-gradient(circle at top left,rgba(45,212,191,.2),transparent 30%),linear-gradient(135deg,#07111f,#0f766e);padding:20px}
.login-card{width:100%;max-width:420px;background:rgba(255,255,255,.94);border-radius:26px;padding:32px;box-shadow:0 24px 80px rgba(2,6,23,.35)}
.logo{height:44px;object-fit:contain}
.kpi-line{display:flex;gap:10px;flex-wrap:wrap}.kpi-line .badge{font-size:12px}
@media(max-width:1000px){.topbar{align-items:flex-start;flex-direction:column}.form-row,.form-row-2,.form-row-3,.grid-2,.grid-3,.grid-4{grid-template-columns:1fr}.brand{min-width:auto}}
</style>
</head>
<body>
{% if login_screen %}
  {{ body|safe }}
{% else %}
  <div class="topbar">
    <div class="brand">
      <strong>{{ app_name }}</strong>
      <span>{{ app_version }} · {{ company.name if company else '' }}</span>
    </div>
    <div class="nav">
      {% if has_perm('dashboard') %}<a href="{{ url_for('index') }}">Dashboard</a>{% endif %}
      {% if has_perm('pos') %}<a href="{{ url_for('pos') }}">POS</a>{% endif %}
      {% if has_perm('quotes') %}<a href="{{ url_for('quotes') }}">Cotizaciones</a>{% endif %}
      {% if has_perm('products') %}<a href="{{ url_for('products') }}">Productos</a>{% endif %}
      {% if has_perm('stock') %}<a href="{{ url_for('stock') }}">Stock Inteligente</a>{% endif %}
      {% if has_perm('purchases') %}<a href="{{ url_for('purchases') }}">Compras</a>{% endif %}
      {% if has_perm('deliveries') %}<a href="{{ url_for('deliveries') }}">Despachos</a>{% endif %}
      {% if has_perm('cash') %}<a href="{{ url_for('cash') }}">Caja</a>{% endif %}
      {% if has_perm('customers') %}<a href="{{ url_for('customers') }}">Clientes</a>{% endif %}
      {% if has_perm('suppliers') %}<a href="{{ url_for('suppliers') }}">Proveedores</a>{% endif %}
      {% if has_perm('dte') %}<a href="{{ url_for('dte') }}">DTE</a>{% endif %}
      {% if has_perm('ai') %}<a href="{{ url_for('ai_assistant') }}">IA</a>{% endif %}
      {% if has_perm('users') %}<a href="{{ url_for('users') }}">Usuarios</a>{% endif %}
      {% if has_perm('audit') %}<a href="{{ url_for('audit') }}">Auditoría</a>{% endif %}
    </div>
    <div class="userbox">
      {{ user.full_name }}<br>
      <a href="{{ url_for('logout') }}" style="color:#93c5fd">Salir</a>
    </div>
  </div>
  <main class="layout">
    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        {% for category, message in messages %}
          <div class="flash {{ category }}">{{ message }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}
    {{ body|safe }}
  </main>
{% endif %}
</body>
</html>
"""


def render_page(title, template, login_screen=False, **ctx):
    body = render_template_string(template, **ctx)
    return render_template_string(
        BASE_HTML,
        title=title,
        app_name=APP_NAME,
        app_version=APP_VERSION,
        body=body,
        login_screen=login_screen,
        user=current_user(),
        company=current_company(),
        has_perm=has_perm,
        is_admin=is_admin(),
        is_superior=is_superior(),
    )


# ============================================================
# INIT
# ============================================================

def ensure_seed_data():
    company = Company.query.first()
    if not company:
        company = Company()
        db.session.add(company)
        db.session.commit()

    if not Branch.query.filter_by(company_id=company.id, code="001").first():
        b1 = Branch(company_id=company.id, code="001", name="Sucursal 1", address=company.address_1)
        b2 = Branch(company_id=company.id, code="002", name="Sucursal 2", address=company.address_2)
        db.session.add_all([b1, b2])
        db.session.commit()
        db.session.add_all([
            Warehouse(company_id=company.id, branch_id=b1.id, code="001", name="Bodega principal"),
            Warehouse(company_id=company.id, branch_id=b2.id, code="002", name="Bodega sucursal 2"),
        ])
        db.session.commit()

    def create_user(username, full_name, password, role, permissions):
        if not User.query.filter_by(username=username).first():
            db.session.add(User(
                company_id=company.id,
                username=username,
                full_name=full_name,
                password_hash=generate_password_hash(password),
                role=role,
                permissions=json.dumps(permissions),
            ))
            db.session.commit()

    all_perms = {k: True for k in PERMISSION_LABELS}
    create_user("gus", "Gustavo Ruz", os.environ.get("GUS_INITIAL_PASSWORD", "gus123"), "superior", all_perms)
    create_user("admin", "Administrador", os.environ.get("ADMIN_INITIAL_PASSWORD", "admin123"), "admin", all_perms)
    create_user("vendedor", "Vendedor", "vendedor123", "vendedor", {"pos": True, "quotes": True, "products": True, "stock": True, "customers": True})
    create_user("bodega", "Bodega", "bodega123", "bodega", {"stock": True, "deliveries": True})


with app.app_context():
    db.create_all()
    ensure_seed_data()


# ============================================================
# AUTH
# ============================================================

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and check_password_hash(user.password_hash, password):
            session["user_id"] = user.id
            write_audit("auth", "login", "user", user.id, "", username)
            return redirect(url_for("index"))
        flash("Usuario o contraseña incorrectos.", "error")

    return render_page("Login", """
    <div class="login-bg">
      <div class="login-card">
        <div style="display:flex;justify-content:center;margin-bottom:18px;">
          <img class="logo" src="{{ url_for('static', filename='ferreteria_san_pedro_logo.jpg') }}" onerror="this.style.display='none'">
        </div>
        <h1 style="margin-bottom:6px;text-align:center;">Ferretería Cloud ERP</h1>
        <p style="text-align:center;margin-bottom:22px;">Plataforma inteligente para retail, stock, ventas y operación.</p>
        {% with messages = get_flashed_messages(with_categories=true) %}
          {% if messages %}
            {% for category, message in messages %}
              <div class="flash {{ category }}">{{ message }}</div>
            {% endfor %}
          {% endif %}
        {% endwith %}
        <form method="post">
          <label>Usuario</label><input name="username" autocomplete="username" required autofocus>
          <div style="height:12px"></div>
          <label>Contraseña</label><input name="password" type="password" autocomplete="current-password" required>
          <div class="actions" style="margin-top:18px;">
            <button class="btn btn-primary" style="width:100%;">Ingresar</button>
          </div>
        </form>
        <p style="text-align:center;font-size:12px;margin-top:18px;">RUZ Technology company</p>
      </div>
    </div>
    """, login_screen=True)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ============================================================
# DASHBOARD
# ============================================================

@app.route("/")
@login_required
@permission_required("dashboard")
def index():
    company = current_company()
    stats = {
        "productos": Product.query.filter_by(company_id=company.id).count(),
        "stock_mov": StockMovement.query.filter_by(company_id=company.id).count(),
        "ventas": Sale.query.filter_by(company_id=company.id).count(),
        "cotizaciones": Quote.query.filter_by(company_id=company.id).count(),
        "despachos": Delivery.query.filter_by(company_id=company.id).count(),
        "clientes": Customer.query.filter_by(company_id=company.id).count(),
    }
    recent_sales = Sale.query.filter_by(company_id=company.id).order_by(Sale.id.desc()).limit(8).all()
    low_stock = db.session.query(Product, StockBalance).join(StockBalance, Product.id == StockBalance.product_id).filter(
        Product.company_id == company.id, StockBalance.available <= 0
    ).limit(10).all()

    return render_page("Dashboard", """
    <h1>Dashboard ejecutivo</h1>
    <p>Base ERP moderna para Ferretería San Pedro: POS, stock inteligente, compras, ventas, despacho, cotización e IA.</p>

    <div class="grid grid-4">
      <div class="stat"><span>Productos</span><strong>{{ stats.productos }}</strong></div>
      <div class="stat"><span>Movimientos stock</span><strong>{{ stats.stock_mov }}</strong></div>
      <div class="stat"><span>Ventas</span><strong>{{ stats.ventas }}</strong></div>
      <div class="stat"><span>Cotizaciones</span><strong>{{ stats.cotizaciones }}</strong></div>
      <div class="stat"><span>Despachos</span><strong>{{ stats.despachos }}</strong></div>
      <div class="stat"><span>Clientes</span><strong>{{ stats.clientes }}</strong></div>
      <div class="stat"><span>IVA</span><strong>{{ iva|percent }}</strong></div>
      <div class="stat"><span>Modo</span><strong style="font-size:16px;">ERP MVP</strong></div>
    </div>

    <div class="grid grid-2" style="margin-top:18px;">
      <div class="card">
        <h2>Ventas recientes</h2>
        <div class="table-wrap"><table class="table">
          <tr><th>Doc.</th><th>Total</th><th>Contribución</th><th>Fecha</th></tr>
          {% for v in recent_sales %}
          <tr><td>{{ v.document_type }} {{ v.document_number }}</td><td>{{ v.total_gross|money }}</td><td>{{ v.contribution|money }}</td><td>{{ v.created_at }}</td></tr>
          {% else %}<tr><td colspan="4" class="muted">Sin ventas.</td></tr>{% endfor %}
        </table></div>
      </div>
      <div class="card">
        <h2>Alertas de stock</h2>
        <div class="table-wrap"><table class="table">
          <tr><th>Código</th><th>Producto</th><th>Bodega</th><th>Disponible</th></tr>
          {% for p,s in low_stock %}
          <tr><td>{{ p.sku }}</td><td>{{ p.description }}</td><td>{{ s.warehouse.name }}</td><td class="num">{{ "%.2f"|format(s.available or 0) }}</td></tr>
          {% else %}<tr><td colspan="4" class="muted">Sin alertas.</td></tr>{% endfor %}
        </table></div>
      </div>
    </div>
    """, stats=stats, recent_sales=recent_sales, low_stock=low_stock, iva=IVA_RATE)


# ============================================================
# PRODUCTOS
# ============================================================

@app.route("/products", methods=["GET", "POST"])
@login_required
@permission_required("products")
def products():
    company = current_company()

    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            sku = request.form.get("sku", "").strip()
            description = request.form.get("description", "").strip()
            if not sku or not description:
                flash("Código y descripción son obligatorios.", "error")
            elif product_by_sku(sku):
                flash("Ese código de producto ya existe.", "error")
            else:
                p = Product(
                    company_id=company.id,
                    sku=sku,
                    barcode=request.form.get("barcode", "").strip(),
                    description=description,
                    category=request.form.get("category", "").strip(),
                    brand=request.form.get("brand", "").strip(),
                    unit=request.form.get("unit", "UN").strip() or "UN",
                    cost_net=parse_float(request.form.get("cost_net")),
                    sale_price_gross=parse_float(request.form.get("sale_price_gross")),
                    min_price_gross=parse_float(request.form.get("min_price_gross")),
                )
                db.session.add(p)
                db.session.commit()
                branch, wh = default_branch_and_warehouse()
                stock_qty = parse_float(request.form.get("stock_initial"))
                if stock_qty:
                    add_stock_movement(p, branch.id, wh.id, "stock_inicial", qty_in=stock_qty, document_type="INI", document_number=f"INI-{p.sku}", cost_net=p.cost_net, notes="Stock inicial producto")
                write_audit("products", "crear", "product", p.id, "", p.sku)
                flash("Producto creado.", "success")
                return redirect(url_for("products"))

        if action == "import":
            file = request.files.get("excel")
            if not file or not file.filename:
                flash("Debes seleccionar un Excel.", "error")
            else:
                result = import_products_excel(file)
                flash(f"Importación productos: {result['total']} filas, {result['created']} creados, {result['updated']} actualizados, {result['errors']} errores.", "success")
                return redirect(url_for("products"))

    q = request.args.get("q", "").strip()
    query = Product.query.filter_by(company_id=company.id)
    if q:
        like = f"%{q}%"
        query = query.filter(db.or_(Product.sku.ilike(like), Product.description.ilike(like), Product.category.ilike(like)))
    rows = query.order_by(Product.sku).limit(500).all()

    return render_page("Productos", """
    <h1>Productos y precios</h1>
    <p>Maestro propio del ERP. Random se usa como referencia histórica; este sistema opera con datos propios.</p>

    <div class="card">
      <h2>Crear producto</h2>
      <form method="post">
        <input type="hidden" name="action" value="create">
        <div class="form-row">
          <div><label>Código SKU *</label><input name="sku" required></div>
          <div><label>Descripción *</label><input name="description" required></div>
          <div><label>Categoría</label><input name="category"></div>
          <div><label>Marca</label><input name="brand"></div>
        </div>
        <div class="form-row" style="margin-top:12px;">
          <div><label>Código barras</label><input name="barcode"></div>
          <div><label>Unidad</label><input name="unit" value="UN"></div>
          <div><label>Costo neto</label><input name="cost_net" inputmode="decimal"></div>
          <div><label>Precio venta bruto</label><input name="sale_price_gross" inputmode="decimal"></div>
        </div>
        <div class="form-row-2" style="margin-top:12px;">
          <div><label>Precio mínimo bruto</label><input name="min_price_gross" inputmode="decimal"></div>
          <div><label>Stock inicial bodega principal</label><input name="stock_initial" inputmode="decimal"></div>
        </div>
        <div class="actions"><button class="btn btn-primary">Guardar producto</button></div>
      </form>
    </div>

    <div class="card">
      <h2>Importar maestro Excel</h2>
      <p>Columnas aceptadas: Código Producto, Descripción, Precio Compra Neto, Precio Venta Bruto, Stock, Activo, Categoría, Marca, Código Barras.</p>
      <form method="post" enctype="multipart/form-data">
        <input type="hidden" name="action" value="import">
        <div class="form-row-2">
          <input type="file" name="excel" accept=".xlsx,.xlsm" required>
          <button class="btn btn-secondary">Importar Excel</button>
        </div>
      </form>
    </div>

    <div class="card">
      <form method="get" class="form-row-2">
        <div><label>Buscar</label><input name="q" value="{{ request.args.get('q','') }}" placeholder="Código, descripción, categoría"></div>
        <div class="actions" style="margin-top:22px;"><button class="btn btn-primary">Filtrar</button><a class="btn btn-secondary" href="{{ url_for('products') }}">Limpiar</a></div>
      </form>
      <div class="table-wrap" style="margin-top:14px;"><table class="table">
        <tr><th>Código</th><th>Descripción</th><th>Categoría</th><th>Costo neto</th><th>Precio bruto</th><th>Activo</th></tr>
        {% for p in rows %}
        <tr><td><b>{{ p.sku }}</b></td><td>{{ p.description }}</td><td>{{ p.category }}</td><td>{{ p.cost_net|money }}</td><td>{{ p.sale_price_gross|money }}</td><td>{% if p.active %}<span class="badge ok">Activo</span>{% else %}<span class="badge bad">Inactivo</span>{% endif %}</td></tr>
        {% else %}<tr><td colspan="6" class="muted">Sin productos.</td></tr>{% endfor %}
      </table></div>
    </div>
    """, rows=rows, request=request)


def import_products_excel(file_storage):
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return {"total": 0, "created": 0, "updated": 0, "errors": 0}

    def norm(v):
        return str(v or "").strip().lower().replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")

    idx = {norm(h): i for i, h in enumerate(headers)}
    def col(*names):
        for n in names:
            if norm(n) in idx:
                return idx[norm(n)]
        return None

    c_sku = col("Código Producto", "Codigo Producto", "Codigo", "SKU")
    c_desc = col("Descripción", "Descripcion", "Producto")
    c_cost = col("Precio Compra Neto", "Costo Neto", "Costo")
    c_price = col("Precio Venta Bruto", "Precio Bruto", "Precio Venta")
    c_stock = col("Stock")
    c_active = col("Activo")
    c_cat = col("Categoria", "Categoría")
    c_brand = col("Marca")
    c_barcode = col("Codigo Barras", "Código Barras", "Barcode")

    company = current_company()
    branch, wh = default_branch_and_warehouse()
    total = created = updated = errors = 0

    for row in rows:
        total += 1
        try:
            sku = str(row[c_sku]).strip() if c_sku is not None and row[c_sku] is not None else ""
            desc = str(row[c_desc]).strip() if c_desc is not None and row[c_desc] is not None else ""
            if not sku or not desc:
                errors += 1
                continue
            p = product_by_sku(sku)
            if not p:
                p = Product(company_id=company.id, sku=sku, description=desc)
                db.session.add(p)
                db.session.flush()
                created += 1
            else:
                updated += 1
            p.description = desc
            p.cost_net = parse_float(row[c_cost] if c_cost is not None else 0)
            p.sale_price_gross = parse_float(row[c_price] if c_price is not None else 0)
            p.category = str(row[c_cat]).strip() if c_cat is not None and row[c_cat] is not None else p.category
            p.brand = str(row[c_brand]).strip() if c_brand is not None and row[c_brand] is not None else p.brand
            p.barcode = str(row[c_barcode]).strip() if c_barcode is not None and row[c_barcode] is not None else p.barcode
            if c_active is not None and row[c_active] is not None:
                p.active = str(row[c_active]).strip().lower() in ("1", "si", "sí", "true", "activo", "activa")
            p.updated_at = now_str()
            db.session.commit()

            if c_stock is not None and row[c_stock] is not None:
                target_stock = parse_float(row[c_stock])
                bal = get_or_create_balance(p.id, branch.id, wh.id)
                diff = target_stock - float(bal.physical or 0)
                if abs(diff) > 0.0001:
                    if diff > 0:
                        add_stock_movement(p, branch.id, wh.id, "ajuste_importacion", qty_in=diff, document_type="IMP", document_number="IMPORT_EXCEL", cost_net=p.cost_net, notes="Ajuste por importación Excel")
                    else:
                        add_stock_movement(p, branch.id, wh.id, "ajuste_importacion", qty_out=abs(diff), document_type="IMP", document_number="IMPORT_EXCEL", cost_net=p.cost_net, notes="Ajuste por importación Excel")
        except Exception:
            db.session.rollback()
            errors += 1

    write_audit("products", "importar_excel", "product", "", "", f"{total} filas")
    return {"total": total, "created": created, "updated": updated, "errors": errors}


# ============================================================
# STOCK INTELIGENTE / KARDEX
# ============================================================

@app.route("/stock", methods=["GET", "POST"])
@login_required
@permission_required("stock")
def stock():
    company = current_company()
    branch, wh = default_branch_and_warehouse()

    if request.method == "POST":
        sku = request.form.get("sku", "").strip()
        product = product_by_sku(sku)
        if not product:
            flash("Producto no encontrado.", "error")
        else:
            movement_type = request.form.get("movement_type", "ajuste")
            qty = parse_float(request.form.get("qty"))
            direction = request.form.get("direction", "in")
            if qty <= 0:
                flash("La cantidad debe ser mayor a cero.", "error")
            else:
                if direction == "in":
                    add_stock_movement(product, branch.id, wh.id, movement_type, qty_in=qty, document_type="AJU", document_number=f"AJU-{now_cl().timestamp():.0f}", cost_net=product.cost_net, notes=request.form.get("notes", ""))
                else:
                    add_stock_movement(product, branch.id, wh.id, movement_type, qty_out=qty, document_type="AJU", document_number=f"AJU-{now_cl().timestamp():.0f}", cost_net=product.cost_net, notes=request.form.get("notes", ""))
                write_audit("stock", "movimiento_manual", "product", product.id, "", f"{movement_type} {qty}")
                flash("Movimiento de stock registrado.", "success")
                return redirect(url_for("stock", q=sku))

    q = request.args.get("q", "").strip()
    selected = None
    movements = []
    balances = []
    products = Product.query.filter_by(company_id=company.id).order_by(Product.sku).limit(500).all()

    if q:
        selected = Product.query.filter_by(company_id=company.id, sku=q).first()
        if not selected:
            selected = Product.query.filter(Product.company_id == company.id, Product.description.ilike(f"%{q}%")).first()
        if selected:
            movements = StockMovement.query.filter_by(company_id=company.id, product_id=selected.id).order_by(StockMovement.created_at.desc(), StockMovement.id.desc()).limit(500).all()
            balances = StockBalance.query.filter_by(company_id=company.id, product_id=selected.id).all()

    return render_page("Stock Inteligente", """
    <h1>Stock Inteligente / Kardex Integrado</h1>
    <p>Inspirado en la Consulta Integrada de Stock de Random: movimientos, stock teórico, auditoría por producto y bodega.</p>

    <div class="card">
      <form method="get" class="form-row-2">
        <div><label>Producto</label><input name="q" list="productos" value="{{ request.args.get('q','') }}" placeholder="Código o descripción"></div>
        <datalist id="productos">{% for p in products %}<option value="{{ p.sku }}">{{ p.description }}</option>{% endfor %}</datalist>
        <div class="actions" style="margin-top:22px;"><button class="btn btn-primary">Consultar Kardex</button></div>
      </form>
    </div>

    {% if selected %}
    <div class="grid grid-4">
      <div class="stat"><span>Producto</span><strong style="font-size:16px;">{{ selected.sku }}</strong></div>
      <div class="stat"><span>Descripción</span><strong style="font-size:15px;">{{ selected.description[:60] }}</strong></div>
      <div class="stat"><span>Costo neto</span><strong>{{ selected.cost_net|money }}</strong></div>
      <div class="stat"><span>Precio bruto</span><strong>{{ selected.sale_price_gross|money }}</strong></div>
    </div>

    <div class="card">
      <h2>Saldo por bodega</h2>
      <table class="table">
        <tr><th>Sucursal</th><th>Bodega</th><th>Stock físico</th><th>Reservado</th><th>Comprometido</th><th>Disponible</th></tr>
        {% for b in balances %}
        <tr><td>{{ b.branch.name }}</td><td>{{ b.warehouse.name }}</td><td class="num">{{ "%.2f"|format(b.physical or 0) }}</td><td class="num">{{ "%.2f"|format(b.reserved or 0) }}</td><td class="num">{{ "%.2f"|format(b.committed or 0) }}</td><td class="num">{{ "%.2f"|format(b.available or 0) }}</td></tr>
        {% else %}<tr><td colspan="6" class="muted">Sin saldo.</td></tr>{% endfor %}
      </table>
    </div>

    <div class="card">
      <h2>Movimiento manual / ajuste controlado</h2>
      <form method="post">
        <input type="hidden" name="sku" value="{{ selected.sku }}">
        <div class="form-row">
          <div><label>Tipo</label><select name="movement_type"><option>ajuste</option><option>inventario</option><option>merma</option><option>traslado</option></select></div>
          <div><label>Dirección</label><select name="direction"><option value="in">Entrada</option><option value="out">Salida</option></select></div>
          <div><label>Cantidad</label><input name="qty" inputmode="decimal" required></div>
          <div><label>Observación</label><input name="notes"></div>
        </div>
        <div class="actions"><button class="btn btn-primary">Registrar movimiento</button></div>
      </form>
    </div>

    <div class="card">
      <h2>Kardex / movimientos</h2>
      <div class="table-wrap"><table class="table">
        <tr><th>Fecha</th><th>Tipo mov.</th><th>Documento</th><th>Entrada</th><th>Salida</th><th>Stock después</th><th>Costo</th><th>Precio</th><th>Notas</th></tr>
        {% for m in movements %}
        <tr><td>{{ m.created_at }}</td><td>{{ m.movement_type }}</td><td>{{ m.document_type }} {{ m.document_number }}</td><td class="num">{{ "%.2f"|format(m.qty_in or 0) }}</td><td class="num">{{ "%.2f"|format(m.qty_out or 0) }}</td><td class="num">{{ "%.2f"|format(m.stock_after or 0) }}</td><td>{{ m.cost_net|money }}</td><td>{{ m.price_gross|money }}</td><td>{{ m.notes }}</td></tr>
        {% else %}<tr><td colspan="9" class="muted">Sin movimientos.</td></tr>{% endfor %}
      </table></div>
    </div>
    {% endif %}
    """, selected=selected, movements=movements, balances=balances, products=products, request=request)


# ============================================================
# POS / VENTAS
# ============================================================

@app.route("/pos", methods=["GET", "POST"])
@login_required
@permission_required("pos")
def pos():
    company = current_company()
    branch, wh = default_branch_and_warehouse()
    products = Product.query.filter_by(company_id=company.id, active=True).order_by(Product.sku).limit(500).all()
    customers = Customer.query.filter_by(company_id=company.id, active=True).order_by(Customer.name).limit(300).all()

    if request.method == "POST":
        doc_type = request.form.get("document_type", "Boleta").strip()
        doc_number = request.form.get("document_number", "").strip()
        if not doc_number:
            flash("Número de documento es obligatorio.", "error")
            return redirect(url_for("pos"))

        # Anti duplicado operacional
        if Sale.query.filter_by(company_id=company.id, document_type=doc_type, document_number=doc_number).first():
            flash(f"ALERTA CRÍTICA: el documento {doc_type} {doc_number} ya existe en ventas. No se permite duplicar.", "error")
            return redirect(url_for("pos"))

        customer_id = request.form.get("customer_id") or None
        sale = Sale(company_id=company.id, branch_id=branch.id, warehouse_id=wh.id, customer_id=customer_id, user_id=current_user().id, document_type=doc_type, document_number=doc_number, payment_method=request.form.get("payment_method", "Efectivo"))
        db.session.add(sale)
        db.session.flush()

        total_gross = cost_total = contribution = 0
        for i in range(1, 11):
            sku = request.form.get(f"sku_{i}", "").strip()
            if not sku:
                continue
            product = product_by_sku(sku)
            if not product:
                continue
            qty = parse_float(request.form.get(f"qty_{i}"), 1)
            price = parse_float(request.form.get(f"price_{i}"), product.sale_price_gross)
            discount = parse_float(request.form.get(f"discount_{i}"), 0)
            calc = calc_line(qty, price, discount, product.cost_net)
            line = SaleLine(
                sale_id=sale.id,
                product_id=product.id,
                sku=product.sku,
                description=product.description,
                qty=qty,
                price_gross=price,
                discount_pct=discount,
                final_price_gross=calc["final_price"],
                cost_net=product.cost_net,
                total_gross=calc["total_gross"],
                contribution=calc["contribution"],
                margin_pct=calc["margin"],
            )
            db.session.add(line)
            total_gross += calc["total_gross"]
            cost_total += product.cost_net * qty
            contribution += calc["contribution"]
            add_stock_movement(product, branch.id, wh.id, "venta", qty_out=qty, document_type=doc_type, document_number=doc_number, price_gross=calc["final_price"], cost_net=product.cost_net, notes="Salida por POS")

        sale.total_gross = total_gross
        sale.total_net = total_gross / (1 + IVA_RATE) if total_gross else 0
        sale.iva = total_gross - sale.total_net
        sale.cost_total_net = cost_total
        sale.contribution = contribution
        sale.margin_pct = (total_gross / (cost_total * (1 + IVA_RATE)) - 1) if cost_total > 0 else 0
        db.session.commit()
        write_audit("pos", "crear_venta", "sale", sale.id, "", f"{doc_type} {doc_number}")
        flash("Venta registrada y stock descontado.", "success")
        return redirect(url_for("sale_detail", sale_id=sale.id))

    return render_page("POS", """
    <h1>POS / Punto de Venta</h1>
    <p>Venta rápida con control de margen, contribución, pagos y salida automática de stock.</p>

    <div class="card">
      <form method="post" id="saleForm">
        <div class="form-row">
          <div><label>Tipo documento</label><select name="document_type"><option>Boleta</option><option>Factura</option><option>Guía</option></select></div>
          <div><label>Número documento *</label><input name="document_number" required></div>
          <div><label>Cliente</label><select name="customer_id"><option value="">Cliente tienda</option>{% for c in customers %}<option value="{{ c.id }}">{{ c.name }}</option>{% endfor %}</select></div>
          <div><label>Medio de pago</label><select name="payment_method"><option>Efectivo</option><option>Tarjeta</option><option>Transferencia</option><option>Mixto</option><option>Crédito interno</option></select></div>
        </div>

        <datalist id="productsList">{% for p in products %}<option value="{{ p.sku }}">{{ p.description }} · {{ p.sale_price_gross|money }}</option>{% endfor %}</datalist>

        <div class="table-wrap" style="margin-top:18px;"><table class="table">
          <tr><th>#</th><th>Código</th><th>Cantidad</th><th>Precio bruto</th><th>Descuento %</th></tr>
          {% for i in range(1,11) %}
          <tr>
            <td>{{ i }}</td>
            <td><input name="sku_{{ i }}" list="productsList"></td>
            <td><input name="qty_{{ i }}" value="1" inputmode="decimal"></td>
            <td><input name="price_{{ i }}" inputmode="decimal" placeholder="auto si vacío"></td>
            <td><input name="discount_{{ i }}" value="0" inputmode="decimal"></td>
          </tr>
          {% endfor %}
        </table></div>
        <div class="actions"><button class="btn btn-primary">Registrar venta</button></div>
      </form>
    </div>
    """, products=products, customers=customers, range=range)


@app.route("/sales/<int:sale_id>")
@login_required
@permission_required("pos")
def sale_detail(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    lines = SaleLine.query.filter_by(sale_id=sale.id).all()
    return render_page("Venta", """
    <h1>Venta {{ sale.document_type }} {{ sale.document_number }}</h1>
    <div class="grid grid-4">
      <div class="stat"><span>Total bruto</span><strong>{{ sale.total_gross|money }}</strong></div>
      <div class="stat"><span>Contribución</span><strong>{{ sale.contribution|money }}</strong></div>
      <div class="stat"><span>Margen</span><strong>{{ sale.margin_pct|percent }}</strong></div>
      <div class="stat"><span>Pago</span><strong>{{ sale.payment_method }}</strong></div>
    </div>
    <div class="card">
      <table class="table">
        <tr><th>Código</th><th>Producto</th><th>Cant.</th><th>Precio final</th><th>Total</th><th>Contribución</th><th>Margen</th></tr>
        {% for l in lines %}
        <tr><td>{{ l.sku }}</td><td>{{ l.description }}</td><td>{{ l.qty }}</td><td>{{ l.final_price_gross|money }}</td><td>{{ l.total_gross|money }}</td><td>{{ l.contribution|money }}</td><td>{{ l.margin_pct|percent }}</td></tr>
        {% endfor %}
      </table>
    </div>
    """, sale=sale, lines=lines)


# ============================================================
# COTIZACIONES
# ============================================================

@app.route("/quotes", methods=["GET", "POST"])
@login_required
@permission_required("quotes")
def quotes():
    company = current_company()
    products = Product.query.filter_by(company_id=company.id, active=True).order_by(Product.sku).limit(500).all()

    if request.method == "POST":
        number = f"COT-{today_str().replace('-', '')}-{Quote.query.count()+1:06d}"
        quote = Quote(company_id=company.id, number=number, customer_name=request.form.get("customer_name", "Cliente"), contact=request.form.get("contact", ""), user_id=current_user().id, valid_days=int(parse_float(request.form.get("valid_days"), 2)))
        db.session.add(quote)
        db.session.flush()
        total = cost_total = contrib = 0
        for i in range(1, 11):
            sku = request.form.get(f"sku_{i}", "").strip()
            if not sku:
                continue
            product = product_by_sku(sku)
            if not product:
                continue
            qty = parse_float(request.form.get(f"qty_{i}"), 1)
            price = parse_float(request.form.get(f"price_{i}"), product.sale_price_gross)
            discount = parse_float(request.form.get(f"discount_{i}"), 0)
            calc = calc_line(qty, price, discount, product.cost_net)
            db.session.add(QuoteLine(
                quote_id=quote.id, product_id=product.id, sku=product.sku, description=product.description,
                qty=qty, price_gross=price, discount_pct=discount, final_price_gross=calc["final_price"],
                cost_net=product.cost_net, total_gross=calc["total_gross"], contribution=calc["contribution"], margin_pct=calc["margin"]
            ))
            total += calc["total_gross"]
            cost_total += product.cost_net * qty
            contrib += calc["contribution"]
        quote.total_gross = total
        quote.total_net = total / (1 + IVA_RATE) if total else 0
        quote.iva = total - quote.total_net
        quote.cost_total_net = cost_total
        quote.contribution = contrib
        quote.margin_pct = (total / (cost_total * (1 + IVA_RATE)) - 1) if cost_total > 0 else 0
        db.session.commit()
        write_audit("quotes", "crear", "quote", quote.id, "", quote.number)
        flash("Cotización creada.", "success")
        return redirect(url_for("quote_detail", quote_id=quote.id))

    recent = Quote.query.filter_by(company_id=company.id).order_by(Quote.id.desc()).limit(20).all()
    return render_page("Cotizaciones", """
    <h1>Cotizaciones</h1>
    <p>Cotización propia con margen, contribución y PDF comercial.</p>
    <div class="card">
      <form method="post">
        <div class="form-row-3">
          <div><label>Cliente</label><input name="customer_name" value="Cliente tienda"></div>
          <div><label>Contacto</label><input name="contact"></div>
          <div><label>Validez días</label><input name="valid_days" value="2"></div>
        </div>
        <datalist id="quoteProducts">{% for p in products %}<option value="{{ p.sku }}">{{ p.description }}</option>{% endfor %}</datalist>
        <div class="table-wrap" style="margin-top:18px;"><table class="table">
          <tr><th>#</th><th>Código</th><th>Cantidad</th><th>Precio bruto</th><th>Descuento %</th></tr>
          {% for i in range(1,11) %}
          <tr><td>{{ i }}</td><td><input name="sku_{{ i }}" list="quoteProducts"></td><td><input name="qty_{{ i }}" value="1"></td><td><input name="price_{{ i }}" placeholder="auto si vacío"></td><td><input name="discount_{{ i }}" value="0"></td></tr>
          {% endfor %}
        </table></div>
        <div class="actions"><button class="btn btn-primary">Crear cotización</button></div>
      </form>
    </div>
    <div class="card">
      <h2>Recientes</h2>
      <table class="table">
        <tr><th>Número</th><th>Cliente</th><th>Total</th><th>Contribución</th><th>Margen</th><th>Acción</th></tr>
        {% for q in recent %}
        <tr><td>{{ q.number }}</td><td>{{ q.customer_name }}</td><td>{{ q.total_gross|money }}</td><td>{{ q.contribution|money }}</td><td>{{ q.margin_pct|percent }}</td><td><a class="btn btn-secondary btn-small" href="{{ url_for('quote_detail', quote_id=q.id) }}">Ver</a> <a class="btn btn-primary btn-small" href="{{ url_for('quote_pdf', quote_id=q.id) }}">PDF</a></td></tr>
        {% else %}<tr><td colspan="6">Sin cotizaciones.</td></tr>{% endfor %}
      </table>
    </div>
    """, products=products, recent=recent, range=range)


@app.route("/quotes/<int:quote_id>")
@login_required
@permission_required("quotes")
def quote_detail(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    lines = QuoteLine.query.filter_by(quote_id=quote.id).all()
    return render_page("Cotización", """
    <h1>{{ quote.number }}</h1>
    <div class="actions"><a class="btn btn-primary" href="{{ url_for('quote_pdf', quote_id=quote.id) }}">Generar PDF</a><a class="btn btn-secondary" href="{{ url_for('quotes') }}">Volver</a></div>
    <div class="grid grid-4">
      <div class="stat"><span>Total bruto</span><strong>{{ quote.total_gross|money }}</strong></div>
      <div class="stat"><span>Costo total bruto</span><strong>{{ (quote.cost_total_net*(1+iva))|money }}</strong></div>
      <div class="stat"><span>Contribución</span><strong>{{ quote.contribution|money }}</strong></div>
      <div class="stat"><span>Margen</span><strong>{{ quote.margin_pct|percent }}</strong></div>
    </div>
    <div class="card"><table class="table">
      <tr><th>Código</th><th>Producto</th><th>Cant.</th><th>Precio</th><th>Total</th><th>Contrib.</th><th>Margen</th></tr>
      {% for l in lines %}
      <tr><td>{{ l.sku }}</td><td>{{ l.description }}</td><td>{{ l.qty }}</td><td>{{ l.final_price_gross|money }}</td><td>{{ l.total_gross|money }}</td><td>{{ l.contribution|money }}</td><td>{{ l.margin_pct|percent }}</td></tr>
      {% endfor %}
    </table></div>
    """, quote=quote, lines=lines, iva=IVA_RATE)


@app.route("/quotes/<int:quote_id>/pdf")
@login_required
@permission_required("quotes")
def quote_pdf(quote_id):
    quote = Quote.query.get_or_404(quote_id)
    lines = QuoteLine.query.filter_by(quote_id=quote.id).all()
    company = current_company()

    try:
        from xml.sax.saxutils import escape as xml_escape
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    except Exception as exc:
        flash(f"No se pudo generar PDF: {exc}", "error")
        return redirect(url_for("quote_detail", quote_id=quote.id))

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.2*cm, rightMargin=1.2*cm, topMargin=1.1*cm, bottomMargin=1.1*cm)
    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#334155"))
    title = ParagraphStyle("title", parent=styles["Title"], fontSize=18, leading=22, textColor=colors.HexColor("#0f172a"), alignment=0)
    story = []
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "ferreteria_san_pedro_logo.jpg")
    logo = Image(logo_path, width=5*cm, height=1.58*cm) if os.path.exists(logo_path) else Paragraph(company.name, title)
    block = [Paragraph(f"<b>{company.name}</b>", title), Paragraph(f"{company.legal_name} · R.U.T. {company.rut}", small), Paragraph("Cotización comercial", small)]
    story.append(Table([[logo, block]], colWidths=[5.7*cm, 12*cm]))
    story.append(Spacer(1, 8))
    story.append(Table([[
        Paragraph(f"<b>Visítanos en:</b><br/>{company.address_1}<br/>{company.address_2}", small),
        Paragraph(f"<b>WhatsApp:</b><br/>{company.whatsapp_1}<br/>{company.whatsapp_2}", small)
    ]], colWidths=[11.8*cm, 5.9*cm], style=[("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#f8fafc")),("BOX",(0,0),(-1,-1),0.3,colors.HexColor("#dbeafe")),("VALIGN",(0,0),(-1,-1),"TOP")]))
    story.append(Spacer(1, 10))
    story.append(Paragraph("COTIZACIÓN", ParagraphStyle("h", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#0f766e"))))

    header = [["N°", quote.number, "Fecha", quote.created_at], ["Cliente", quote.customer_name or "", "Contacto", quote.contact or ""], ["Vendedor", current_user().full_name if current_user() else "", "Validez", f"{quote.valid_days} días"]]
    story.append(Table(header, colWidths=[3*cm, 6.1*cm, 2.5*cm, 6.0*cm], style=[("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#cbd5e1")),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#f1f5f9")),("BACKGROUND",(2,0),(2,-1),colors.HexColor("#f1f5f9")),("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8)]))
    story.append(Spacer(1, 12))

    data = [["Código", "Descripción", "Cantidad", "Precio Unit.", "Valor"]]
    for l in lines:
        data.append([l.sku, Paragraph(xml_escape(l.description or ""), small), f"{l.qty:.2f}".replace(".", ","), money(l.final_price_gross), money(l.total_gross)])
    table = Table(data, colWidths=[3*cm, 8.4*cm, 2*cm, 2.8*cm, 2.8*cm], repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0f766e")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#d1d5db")),("FONTSIZE",(0,0),(-1,-1),8),("VALIGN",(0,0),(-1,-1),"TOP"),("ALIGN",(2,1),(-1,-1),"RIGHT")]))
    story.append(table)
    story.append(Spacer(1, 12))
    totals = Table([["Neto", money(quote.total_net)], [f"IVA {IVA_RATE*100:.0f}%", money(quote.iva)], ["Total", money(quote.total_gross)]], colWidths=[4*cm,4*cm], hAlign="RIGHT")
    totals.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#cbd5e1")),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#f1f5f9")),("BACKGROUND",(0,2),(-1,2),colors.HexColor("#dcfce7")),("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("ALIGN",(1,0),(1,-1),"RIGHT")]))
    story.append(totals)
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"CONDICIONES COMERCIALES<br/>1.- Validez de la oferta: {quote.valid_days} días.<br/>2.- Plazo de entrega: sujeto a disponibilidad de stock.", small))
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"{quote.number}.pdf", mimetype="application/pdf")


# ============================================================
# COMPRAS, CLIENTES, PROVEEDORES, CAJA, DESPACHOS
# ============================================================

@app.route("/customers", methods=["GET", "POST"])
@login_required
@permission_required("customers")
def customers():
    company = current_company()
    if request.method == "POST":
        c = Customer(company_id=company.id, name=request.form.get("name", "").strip(), rut=request.form.get("rut","").strip(), phone=request.form.get("phone","").strip(), email=request.form.get("email","").strip(), address=request.form.get("address","").strip())
        if not c.name:
            flash("Nombre es obligatorio.", "error")
        else:
            db.session.add(c); db.session.commit(); flash("Cliente creado.", "success"); return redirect(url_for("customers"))
    rows = Customer.query.filter_by(company_id=company.id).order_by(Customer.name).limit(500).all()
    return render_page("Clientes", """
    <h1>Clientes / CRM</h1>
    <div class="card"><form method="post" class="form-row"><div><label>Nombre</label><input name="name" required></div><div><label>RUT</label><input name="rut"></div><div><label>Teléfono</label><input name="phone"></div><div><label>Email</label><input name="email"></div><div style="grid-column:span 4;"><label>Dirección</label><input name="address"></div><div class="actions"><button class="btn btn-primary">Crear cliente</button></div></form></div>
    <div class="card"><table class="table"><tr><th>Nombre</th><th>RUT</th><th>Teléfono</th><th>Email</th><th>Dirección</th></tr>{% for c in rows %}<tr><td>{{ c.name }}</td><td>{{ c.rut }}</td><td>{{ c.phone }}</td><td>{{ c.email }}</td><td>{{ c.address }}</td></tr>{% endfor %}</table></div>
    """, rows=rows)


@app.route("/suppliers", methods=["GET", "POST"])
@login_required
@permission_required("suppliers")
def suppliers():
    company = current_company()
    if request.method == "POST":
        s = Supplier(company_id=company.id, name=request.form.get("name","").strip(), rut=request.form.get("rut","").strip(), phone=request.form.get("phone","").strip(), email=request.form.get("email","").strip(), address=request.form.get("address","").strip())
        if not s.name: flash("Nombre es obligatorio.", "error")
        else: db.session.add(s); db.session.commit(); flash("Proveedor creado.", "success"); return redirect(url_for("suppliers"))
    rows = Supplier.query.filter_by(company_id=company.id).order_by(Supplier.name).limit(500).all()
    return render_page("Proveedores", """
    <h1>Proveedores</h1>
    <div class="card"><form method="post" class="form-row"><div><label>Nombre</label><input name="name" required></div><div><label>RUT</label><input name="rut"></div><div><label>Teléfono</label><input name="phone"></div><div><label>Email</label><input name="email"></div><div style="grid-column:span 4;"><label>Dirección</label><input name="address"></div><div class="actions"><button class="btn btn-primary">Crear proveedor</button></div></form></div>
    <div class="card"><table class="table"><tr><th>Nombre</th><th>RUT</th><th>Teléfono</th><th>Email</th><th>Dirección</th></tr>{% for s in rows %}<tr><td>{{ s.name }}</td><td>{{ s.rut }}</td><td>{{ s.phone }}</td><td>{{ s.email }}</td><td>{{ s.address }}</td></tr>{% endfor %}</table></div>
    """, rows=rows)


@app.route("/purchases", methods=["GET", "POST"])
@login_required
@permission_required("purchases")
def purchases():
    company = current_company()
    branch, wh = default_branch_and_warehouse()
    suppliers_list = Supplier.query.filter_by(company_id=company.id, active=True).order_by(Supplier.name).limit(300).all()
    products_list = Product.query.filter_by(company_id=company.id, active=True).order_by(Product.sku).limit(500).all()
    if request.method == "POST":
        purchase = Purchase(company_id=company.id, supplier_id=request.form.get("supplier_id") or None, branch_id=branch.id, warehouse_id=wh.id, document_type=request.form.get("document_type","Factura compra"), document_number=request.form.get("document_number","").strip())
        db.session.add(purchase); db.session.flush()
        total_net = 0
        for i in range(1, 8):
            sku = request.form.get(f"sku_{i}", "").strip()
            if not sku: continue
            product = product_by_sku(sku)
            if not product: continue
            qty = parse_float(request.form.get(f"qty_{i}"), 1)
            cost = parse_float(request.form.get(f"cost_{i}"), product.cost_net)
            db.session.add(PurchaseLine(purchase_id=purchase.id, product_id=product.id, sku=product.sku, description=product.description, qty=qty, cost_net=cost, total_net=cost*qty))
            product.cost_net = cost
            total_net += cost * qty
            add_stock_movement(product, branch.id, wh.id, "compra_recepcion", qty_in=qty, document_type=purchase.document_type, document_number=purchase.document_number, cost_net=cost, notes="Entrada por compra")
        purchase.total_net = total_net; purchase.iva = total_net*IVA_RATE; purchase.total_gross = total_net*(1+IVA_RATE)
        db.session.commit(); flash("Compra/recepción registrada.", "success"); return redirect(url_for("purchases"))
    recent = Purchase.query.filter_by(company_id=company.id).order_by(Purchase.id.desc()).limit(20).all()
    return render_page("Compras", """
    <h1>Compras y recepción</h1>
    <div class="card"><form method="post">
      <div class="form-row"><div><label>Proveedor</label><select name="supplier_id"><option value="">Sin proveedor</option>{% for s in suppliers_list %}<option value="{{ s.id }}">{{ s.name }}</option>{% endfor %}</select></div><div><label>Tipo doc.</label><input name="document_type" value="Factura compra"></div><div><label>Número doc.</label><input name="document_number" required></div></div>
      <datalist id="plist">{% for p in products_list %}<option value="{{ p.sku }}">{{ p.description }}</option>{% endfor %}</datalist>
      <table class="table" style="margin-top:14px;"><tr><th>#</th><th>Código</th><th>Cantidad</th><th>Costo neto</th></tr>{% for i in range(1,8) %}<tr><td>{{ i }}</td><td><input name="sku_{{ i }}" list="plist"></td><td><input name="qty_{{ i }}" value="1"></td><td><input name="cost_{{ i }}"></td></tr>{% endfor %}</table>
      <div class="actions"><button class="btn btn-primary">Registrar recepción</button></div>
    </form></div>
    <div class="card"><h2>Últimas compras</h2><table class="table"><tr><th>Doc.</th><th>Neto</th><th>Total</th><th>Fecha</th></tr>{% for p in recent %}<tr><td>{{ p.document_type }} {{ p.document_number }}</td><td>{{ p.total_net|money }}</td><td>{{ p.total_gross|money }}</td><td>{{ p.created_at }}</td></tr>{% endfor %}</table></div>
    """, suppliers_list=suppliers_list, products_list=products_list, recent=recent, range=range)


@app.route("/cash", methods=["GET", "POST"])
@login_required
@permission_required("cash")
def cash():
    company = current_company()
    branch, wh = default_branch_and_warehouse()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "open":
            db.session.add(CashSession(company_id=company.id, branch_id=branch.id, user_id=current_user().id, opening_amount=parse_float(request.form.get("opening_amount"))))
            db.session.commit(); flash("Caja abierta.", "success")
        elif action == "close":
            cid = request.form.get("cash_id")
            cs = CashSession.query.get(cid)
            if cs:
                cs.status = "Cerrada"; cs.closing_amount = parse_float(request.form.get("closing_amount")); cs.closed_at = now_str(); db.session.commit(); flash("Caja cerrada.", "success")
    sessions = CashSession.query.filter_by(company_id=company.id).order_by(CashSession.id.desc()).limit(50).all()
    return render_page("Caja", """
    <h1>Caja y turnos</h1>
    <div class="grid grid-2">
      <div class="card"><h2>Abrir caja</h2><form method="post"><input type="hidden" name="action" value="open"><label>Monto inicial</label><input name="opening_amount"><div class="actions"><button class="btn btn-primary">Abrir</button></div></form></div>
      <div class="card"><h2>Cerrar caja</h2><form method="post"><input type="hidden" name="action" value="close"><label>ID caja</label><input name="cash_id"><label>Monto cierre</label><input name="closing_amount"><div class="actions"><button class="btn btn-secondary">Cerrar</button></div></form></div>
    </div>
    <div class="card"><table class="table"><tr><th>ID</th><th>Estado</th><th>Apertura</th><th>Cierre</th><th>Inicial</th><th>Final</th></tr>{% for c in sessions %}<tr><td>{{ c.id }}</td><td>{{ c.status }}</td><td>{{ c.opened_at }}</td><td>{{ c.closed_at }}</td><td>{{ c.opening_amount|money }}</td><td>{{ c.closing_amount|money }}</td></tr>{% endfor %}</table></div>
    """, sessions=sessions)


@app.route("/deliveries", methods=["GET", "POST"])
@login_required
@permission_required("deliveries")
def deliveries():
    company = current_company()
    if request.method == "POST":
        doc_type = request.form.get("document_type", "Factura").strip()
        doc_number = request.form.get("document_number", "").strip()
        if not doc_number:
            flash("Número documento es obligatorio.", "error")
        elif Delivery.query.filter_by(company_id=company.id, document_type=doc_type, document_number=doc_number).first():
            flash(f"ALERTA CRÍTICA: el documento {doc_type} {doc_number} ya fue registrado en despacho. No se permite duplicar entrega.", "error")
        else:
            d = Delivery(company_id=company.id, document_type=doc_type, document_number=doc_number, status=request.form.get("status","Pendiente"), vehicle_plate=request.form.get("vehicle_plate","").upper(), driver=request.form.get("driver",""), helper=request.form.get("helper",""), notes=request.form.get("notes",""), created_by=current_user().id)
            db.session.add(d); db.session.commit(); write_audit("deliveries", "crear", "delivery", d.id, "", f"{doc_type} {doc_number}"); flash("Despacho registrado.", "success"); return redirect(url_for("deliveries"))
    rows = Delivery.query.filter_by(company_id=company.id).order_by(Delivery.id.desc()).limit(100).all()
    return render_page("Despachos", """
    <h1>Despachos / Bodega</h1>
    <p>Control anti-duplicado por tipo y número de documento. Futura evolución: despacho por línea y entrega parcial.</p>
    <div class="card"><form method="post"><div class="form-row"><div><label>Tipo documento</label><select name="document_type"><option>Factura</option><option>Boleta</option><option>Guía</option></select></div><div><label>Número *</label><input name="document_number" required></div><div><label>Estado</label><select name="status"><option>Pendiente</option><option>Entregado</option></select></div><div><label>Patente</label><input name="vehicle_plate"></div></div><div class="form-row-3" style="margin-top:12px;"><div><label>Chofer</label><input name="driver"></div><div><label>Pioneta</label><input name="helper"></div><div><label>Observación</label><input name="notes"></div></div><div class="actions"><button class="btn btn-primary">Guardar despacho</button></div></form></div>
    <div class="card"><table class="table"><tr><th>ID</th><th>Documento</th><th>Estado</th><th>Patente</th><th>Chofer</th><th>Pioneta</th><th>Fecha</th></tr>{% for d in rows %}<tr><td>{{ d.id }}</td><td>{{ d.document_type }} {{ d.document_number }}</td><td>{% if d.status=='Entregado' %}<span class="badge ok">Entregado</span>{% else %}<span class="badge warn">{{ d.status }}</span>{% endif %}</td><td>{{ d.vehicle_plate }}</td><td>{{ d.driver }}</td><td>{{ d.helper }}</td><td>{{ d.created_at }}</td></tr>{% endfor %}</table></div>
    """, rows=rows)


# ============================================================
# IA, DTE, USUARIOS, AUDITORÍA
# ============================================================

@app.route("/dte")
@login_required
@permission_required("dte")
def dte():
    return render_page("Documentos DTE", """
    <h1>Documentos DTE</h1>
    <div class="card">
      <h2>Integración tributaria</h2>
      <p>Diseñado para integrarse con Facturación.cl u otro proveedor DTE. Facturación.cl no será fuente de productos/stock; será motor de documentos tributarios.</p>
      <ul>
        <li>Obtener PDF / Link por folio.</li>
        <li>Emitir factura, boleta, guía o nota de crédito en etapa futura.</li>
        <li>Leer XML si se contrata getXMLDte.</li>
        <li>Cruzar guías/facturas con despacho y stock.</li>
      </ul>
    </div>
    """)


@app.route("/ai", methods=["GET", "POST"])
@login_required
@permission_required("ai")
def ai_assistant():
    answer = ""
    if request.method == "POST":
        question = request.form.get("question", "").strip()
        context = f"ERP Ferretería San Pedro. Productos: {Product.query.count()}, ventas: {Sale.query.count()}, movimientos stock: {StockMovement.query.count()}."
        if os.environ.get("OPENAI_API_KEY"):
            try:
                from openai import OpenAI
                client = OpenAI()
                response = client.responses.create(
                    model=os.environ.get("OPENAI_MODEL", "gpt-5.4-mini"),
                    input=f"Eres Elias, asistente gerencial de un ERP retail. Contexto: {context}\nPregunta: {question}",
                    max_output_tokens=900,
                )
                answer = getattr(response, "output_text", str(response))
            except Exception as exc:
                answer = f"No se pudo consultar IA: {exc}"
        else:
            answer = "OPENAI_API_KEY no está configurada. La capa IA quedará operativa al agregar la variable de entorno en Render."
    return render_page("IA Elias", """
    <h1>IA Elias / Asistente gerencial</h1>
    <div class="card"><form method="post"><label>Pregunta</label><textarea name="question" placeholder="Ej: ¿Qué productos tienen bajo stock? ¿Qué debería comprar?"></textarea><div class="actions"><button class="btn btn-primary">Consultar</button></div></form></div>
    {% if answer %}<div class="card"><h2>Respuesta</h2><p style="white-space:pre-wrap;">{{ answer }}</p></div>{% endif %}
    """, answer=answer)


@app.route("/users", methods=["GET", "POST"])
@login_required
@permission_required("users")
def users():
    company = current_company()
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","").strip()
        if not username or not password:
            flash("Usuario y contraseña son obligatorios.", "error")
        elif User.query.filter_by(username=username).first():
            flash("Usuario ya existe.", "error")
        else:
            perms = {p: True for p in request.form.getlist("permissions")}
            u = User(company_id=company.id, username=username, full_name=request.form.get("full_name", username), password_hash=generate_password_hash(password), role=request.form.get("role", "operador"), permissions=json.dumps(perms))
            db.session.add(u); db.session.commit(); flash("Usuario creado.", "success")
    rows = User.query.order_by(User.id).all()
    return render_page("Usuarios", """
    <h1>Usuarios y permisos</h1>
    <div class="card"><form method="post"><div class="form-row"><div><label>Usuario</label><input name="username"></div><div><label>Nombre</label><input name="full_name"></div><div><label>Clave</label><input name="password"></div><div><label>Rol</label><select name="role"><option>operador</option><option>admin</option><option>gerente</option><option>vendedor</option><option>bodega</option></select></div></div><h3>Permisos</h3><div class="grid grid-4">{% for k,l in perms.items() %}<label><input type="checkbox" name="permissions" value="{{ k }}"> {{ l }}</label>{% endfor %}</div><div class="actions"><button class="btn btn-primary">Crear usuario</button></div></form></div>
    <div class="card"><table class="table"><tr><th>ID</th><th>Usuario</th><th>Nombre</th><th>Rol</th><th>Superior</th><th>Activo</th><th>Permisos</th></tr>{% for u in rows %}<tr><td>{{ u.id }}</td><td>{{ u.username }}</td><td>{{ u.full_name }}</td><td>{{ u.role }}</td><td>{% if u.username.lower()==superior %}<span class="badge warn">Superior</span>{% endif %}</td><td>{{ u.is_active }}</td><td style="max-width:380px;white-space:pre-wrap;">{{ u.permissions }}</td></tr>{% endfor %}</table></div>
    """, rows=rows, perms=PERMISSION_LABELS, superior=SUPERIOR_USERNAME)


@app.route("/audit")
@login_required
@permission_required("audit")
def audit():
    rows = AuditLog.query.order_by(AuditLog.id.desc()).limit(300).all()
    return render_page("Auditoría", """
    <h1>Auditoría</h1>
    <div class="card"><table class="table"><tr><th>Fecha</th><th>Usuario</th><th>Módulo</th><th>Acción</th><th>Entidad</th><th>Anterior</th><th>Nuevo</th></tr>{% for a in rows %}<tr><td>{{ a.created_at }}</td><td>{{ a.username }}</td><td>{{ a.module }}</td><td>{{ a.action }}</td><td>{{ a.entity }} {{ a.entity_id }}</td><td>{{ a.old_value }}</td><td>{{ a.new_value }}</td></tr>{% endfor %}</table></div>
    """, rows=rows)


@app.route("/health")
def health():
    return {"status": "ok", "app": APP_NAME, "version": APP_VERSION, "time": now_str()}


if __name__ == "__main__":
    app.run(debug=True)
